"""
Module for exporting data to Excel using lightweight opnpyxl.
Does not format Excel data but is lighter than the whole pandas module

Functions:
  -  `exc_export(data, file_name='output')`:
    Writes a list of dictionaries, each containing the information about available boats
    into an excel file
"""

from openpyxl import Workbook

def exc_export(data, file_name='output'):
    """
    Writes a list of dictionaries, each containing the information about available boats
    into an Excel file

    Args:
        data (list): A list of dictionaries, each containing
        the information about available boats.
        file_name (str, optional): String containing the name
        with which the Excel document should be saved.


    Returns:
        None
    """

    # Create a workbook and add a worksheet
    work_book = Workbook()
    work_sheet = work_book.active

    # Add column headers
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, 1):
        work_sheet.cell(row=1, column=col_num, value=header)

        # Add data to the worksheet
    for row_num, row_data in enumerate(data, 2):
        for col_num, key in enumerate(headers, 1):
            work_sheet.cell(row=row_num, column=col_num, value=row_data[key])

    # Save the workbook
    work_book.save(file_name + '.xlsx')
